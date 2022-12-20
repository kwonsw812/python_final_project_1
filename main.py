import os
import re
import openpyxl
import natsort
import zipfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication


count = 0
loop_count = 0

file_path = input("파일 경로를 복사 후 붙여넣으시오. : ")
file_extension = input("변경할 확장자명을 .을 포함해 입력하시오 {ex).jpg} : ")

file_names = os.listdir(file_path)
file_names = natsort.natsorted(file_names)
print(file_names)

sel = input("바꿀 방식을 선택하시오.\n1.입력한 숫자부터 + 1\n"
            "2.문자열 + 입력한 숫자부터 + 1\n3.파일에서 가져오기\n")

if sel == "1" :
    count = input("시작할 숫자를 입력하세요 : ")
    i = int(count)
    for file_name in file_names:
        print(file_path)
        old = file_path+"\\"+file_name
        dst = str(i) + file_extension
        new = file_path+"\\"+dst
        print(dst)
        os.rename(old, new)
        i += 1

elif sel == "2" :
    change = input("문자열을 입력하세요 :")
    count = input("시작할 숫자를 입력하세요 : ")
    i = int(count)
    for file_name in file_names:
        old = file_path+"\\"+file_name
        dst = change + str(i) + file_extension
        new = file_path+"\\"+dst
        os.rename(old, new)
        i += 1

elif sel == "3" :
    file_location = input("사용할 파일의 경로를 복사 후 붙여넣으시오. : ")
    files = os.listdir(file_location)

    for file in files :
        if ".xlsx" or ".xls" in file :
            count += 1

    if count == 1 :
        for file in files :
            if ".xls" in file:
                path = file_location + '/' + file
                path = path.replace('\\', '/')
                wb = openpyxl.load_workbook(path)
                ws = wb.active

                sel_cell = input("사용할 행 또는 열을 입력하시오. : ")
                num = re.sub(r'[^0-9]', '', sel_cell)

                if "행" in sel_cell:
                    if ws.max_column <= len(file_names):
                        i = 0
                        for file_name in file_names:
                            if i <= ws.max_row:
                                print(file_name)
                                old = file_path + "\\" + file_name
                                dst = str(ws.cell(int(num), i + 1).value) + file_extension
                                new = file_path + "\\" + dst
                                print(dst)
                                os.rename(old, new)
                                i += 1

                    else:
                        i = 0
                        for file_name in file_names:
                            if i <= len(file_names):
                                print(file_name)
                                old = file_path + "\\" + file_name
                                dst = str(ws.cell(i + 1, int(num)).value) + file_extension
                                new = file_path + "\\" + dst
                                print(dst)
                                os.rename(old, new)
                                i += 1

                if "열" in sel_cell:
                    if ws.max_row <= len(file_names):
                        i = 0
                        for file_name in file_names:
                            if i <= ws.max_row:
                                print(file_name)
                                old = file_path + "\\" + file_name
                                dst = str(ws.cell(i + 1, int(num)).value) + file_extension
                                new = file_path + "\\" + dst
                                print(dst)
                                os.rename(old, new)
                                i += 1

                    else:
                        i = 0
                        for file_name in file_names:
                            if i <= len(file_names):
                                print(file_name)
                                old = file_path + "\\" + file_name
                                dst = str(ws.cell(i + 1, int(num)).value) + file_extension
                                new = file_path + "\\" + dst
                                print(dst)
                                os.rename(old, new)
                                i += 1

    else :
        for i, file_name in enumerate(files):
            print(i, file_name)

        file_sel = int(input("사용할 파일의 번호를 입력하세요.\n"))

        if ".xls" in files[file_sel]:
            loop_count += 1
            path = file_location + '/' + files[file_sel]
            path = path.replace('\\', '/')
            print(path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            sel_cell = input("사용할 행 또는 열을 입력하시오. : ")
            num = re.sub(r'[^0-9]', '', sel_cell)
            print(num)

            if "행" in sel_cell:
                if ws.max_column <= len(file_names):
                    i = 0
                    for file_name in file_names:
                        if i <= ws.max_row:
                            print(file_name)
                            old = file_path + "\\" + file_name
                            dst = str(ws.cell(int(num), i + 1).value) + file_extension
                            new = file_path + "\\" + dst
                            print(dst)
                            os.rename(old, new)
                            i += 1

                else:
                    i = 0
                    for file_name in file_names:
                        if i <= len(file_names):
                            print(file_name)
                            old = file_path + "\\" + file_name
                            dst = str(ws.cell(i + 1, int(num)).value) + file_extension
                            new = file_path + "\\" + dst
                            print(dst)
                            os.rename(old, new)
                            i += 1

            if "열" in sel_cell:
                if ws.max_row <= len(file_names):
                    i = 0
                    for file_name in file_names:
                        if i <= ws.max_row:
                            print(file_name)
                            old = file_path + "\\" + file_name
                            dst = str(ws.cell(i + 1, int(num)).value) + file_extension
                            new = file_path + "\\" + dst
                            print(dst)
                            os.rename(old, new)
                            i += 1

                else:
                    i = 0
                    for file_name in file_names:
                        if i <= len(file_names):
                            print(file_name)
                            old = file_path + "\\" + file_name
                            dst = str(ws.cell(i + 1, int(num)).value) + file_extension
                            new = file_path + "\\" + dst
                            print(dst)
                            os.rename(old, new)
                            i += 1


else :
    print("1, 2, 3 중에 하나의 숫자를 선택해 주세요.")

zip_file = zipfile.ZipFile(file_path + "/output.zip", "w")
for (path, dir, files_in) in os.walk(file_path) :
    for file in files_in :
        if file.endswith(file_extension) :
            zip_file.write(os.path.join(path, file), compress_type = zipfile.ZIP_DEFLATED)

zip_file.close()

email_sel = int(input("이메일을 보내시겠습니까?\n1. 예 2. 아니오\n"))

if email_sel == 1 :
    my_email = input("자신의 이메일을 입력하세요 : ")
    my_pw = input("자신의 이메일 비밀번호를 입력하세요 : ")
    subject = input("이메일 제목을 입력하세요 : ")
    content = input("보낼 내용을 입력하세요 : ")
    to_email = input("보낼 이메일 주소를 입력하세요 : ")

    msg = MIMEMultipart()
    msg["FROM"] = my_email
    msg["SUBJECT"] = subject
    msg["TO"] = to_email

    content_part = MIMEText(content, "plain")
    msg.attach(content_part)

    email_file = 'D:/123' + '/' + 'output.zip'

    with open(email_file, 'rb') as file:
        msg.attach(MIMEApplication(file.read(), Name='output.zip'))

    smtp = smtplib.SMTP('smtp.naver.com', 587)
    smtp.ehlo()
    smtp.starttls()
    smtp.login(my_email, my_pw)

    smtp.sendmail(my_email, to_email, msg.as_string())

    smtp.quit()

else :
    print("파일 변환을 마칩니다.")